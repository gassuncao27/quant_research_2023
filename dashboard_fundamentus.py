from pymongo import MongoClient
from datetime import datetime, timedelta
from unidecode import unidecode
from dotenv import load_dotenv

import openpyxl

# from openpyxl.styles import Font
# from openpyxl.styles import PatternFill
# from openpyxl import load_workbook

import pandas as pd

import numpy as np
import os
import re

# Acessando o banco MongoDB Compass
cliente = MongoClient(
    "ddns.cartor.com.br",
    port=57017,
    username='cartor',
    password='Cartor1212',
    authSource='admin',
    authMechanism='DEFAULT'
)

db = cliente.CVM
collection = 'Empresas_fin6'

load_dotenv()   
stocks_path = os.getenv('AMBIENTE_MAC')+'stocks_data.parquet'
df_stocks = pd.read_parquet(stocks_path); df_stocks.fillna(0, inplace=True)

# abertura template excel 
arquivo = openpyxl.load_workbook('estquant_fundamentus.xlsx')
planilha = arquivo['Planilha1']

def fGet_adjclosing_price(df, date, ticker):
    if isinstance(date, str):
        date = pd.to_datetime(date)
    filtered_df = df[(df['data'] == date) & (df['ticker'] == ticker)]
    if not filtered_df.empty:
        return filtered_df['preco_fechamento_ajustado'].iloc[0]
    else:
        return None
    
def fEncontrar_close(df, stock_name, fdata_divulgacao):
    try:
        filtro = (df['Stock'] == stock_name) & (df['Date'] == fdata_divulgacao)
        close_value = df.loc[filtro, 'Close'].values[0]
        return close_value
    except IndexError:
        try:
            filtro = (df['Stock'] == stock_name) & (df['Date'] == (fdata_divulgacao+timedelta(days=1)))
            close_value = df.loc[filtro, 'Close'].values[0]
            return close_value            
        except IndexError:        
            try:
                filtro = (df['Stock'] == stock_name) & (df['Date'] == (fdata_divulgacao+timedelta(days=2)))
                close_value = df.loc[filtro, 'Close'].values[0]
                return close_value
            except IndexError:
                try:
                    filtro = (df['Stock'] == stock_name) & (df['Date'] == (fdata_divulgacao+timedelta(days=2)))
                    close_value = df.loc[filtro, 'Close'].values[0]
                    return close_value
                except IndexError:                                                                
                    return None 
    
def fConverterstr_data(fData_hora_string):
    data_hora = datetime.strptime(fData_hora_string, "%d/%m/%Y %H:%M:%S")
    data = data_hora.date()
    return data    

def fExtrair_hora(fData_hora_str):
    data_hora = datetime.strptime(fData_hora_str, "%d/%m/%Y %H:%M:%S")
    horas = data_hora.hour
    if horas > 17:
        aumentar_dias = 1
    else:
        aumentar_dias = 0
    return aumentar_dias

def finddoc_byequity(fDb, fCollection_name, fEquity): # acha o documento da empresa pelo ticker da bolsa 
    count = 0
    while count != 1:
        collection = fDb[fCollection_name]
        documents = collection.find({'equity': fEquity})
        count = collection.count_documents({'equity': fEquity})
        if count > 1:
            print('Lista encontrada: ')
            print(type(documents))
            for doc in documents:
                print(doc['_id'])
            print('CODIGO EXISTENTE EM MAIS DE 01 DOCUMENTO')
            exit()
        if count == 0:
            print("Não foi encontrada a empresa desejada.")
            return None  # Retorna None se não encontrar a empresa
    return documents

def finddoc_byequity(db, collection_name, b3_code):
    count = 0
    while count != 1:
        collection = db[collection_name]
        documents = collection.find({'equity': b3_code})
        count = collection.count_documents({'equity': b3_code})
        if count > 1:
            print('Lista encontrada: ')
            print(type(documents))
            for doc in documents:
                print(doc['_id'])
            print('CODIGO EXISTENTE EM MAIS DE 01 DOCUMENTO')
            exit()
        if count == 0:
            print("Não foi encontrada a empresa desejada.")
            return None  # Retorna None se não encontrar a empresa
    return documents

def finddoc_byname(fDb, fCollection_name, fContaining_word, fTrimestre=False):
    count = 0
    while count != 1:
        if fTrimestre:
            collection = fDb[fCollection_name]
            regex = re.compile(fContaining_word+fTrimestre, re.IGNORECASE)
            documents = collection.find({'_id': regex})
            count = collection.count_documents({'_id': regex})
        else:
            collection = fDb[fCollection_name]
            regex = re.compile(fContaining_word, re.IGNORECASE)
            documents = collection.find({'_id': regex})
            count = collection.count_documents({'_id': regex})
        if count > 1:
            print(type(documents))
            print('erro aconteceu - checar base')
            for doc in documents:
                print(doc['_id'])
            break
        if count == 0:
            print("Não foi encontrada a empresa desejada.")
            return None
    return documents

def upload_infofin(fData_fin, fDoc): # retorna os trimestres da empresa 
    trimestres_upload = []
    for chave, valor in fData_fin.items():
        if isinstance(chave, str) and chave.startswith('2'):
            trimestres_upload.append(chave)
    trimestres_upload = sorted(trimestres_upload, reverse=True)     
    tipo_balanco, trimestres_upload = balance_type(fDoc, trimestres_upload)
    if tipo_balanco == 'ERROR': exit()
    tipo_balanco2 = re.sub(r'(DF)(Consolidado|Individual)', r'\1 \2', tipo_balanco)
    return tipo_balanco, tipo_balanco2, trimestres_upload

def balance_type(fEmpresa, fTrimestre): # seleciona o tipo de balanco existente
    tipo_balanco = 'ERROR'   
    # print('\n','entrei aki','\n')
    for key in fEmpresa[fTrimestre[0]]:
        # print(key)
        key2 = re.sub(r'(DF)(Consolidado|Individual)', r'\1 \2', key)       
        # print(key2)
        try:
             pvalor = fEmpresa[fTrimestre[-1]][key][key2+' - Demonstração do Resultado']['3.01']['vl_conta']
            #  print(fTrimestre[-1])
            #  print(pvalor)
             if pvalor != 0: 
                 tipo_balanco = key
                 break
        except KeyError:
            # print('\nCia nao possui DF CONSOLIDADO')
            pass
    novalista_trimestres = []
    if tipo_balanco == 'DFConsolidado':
        for trimestre in fTrimestre:
            # print(fEmpresa[trimestre])
            if len(fEmpresa[trimestre].get(tipo_balanco, {})) != 0:
                    novalista_trimestres.append(trimestre) 
        fTrimestre = novalista_trimestres
    # print(f'\nUtilizando balanço: {tipo_balanco}')        
    return tipo_balanco, fTrimestre

def encontrar_conta(fMydict, fConta):
   f_value = 0 
   for key in fMydict:
      string = unidecode(fMydict[key]['ds_conta'].lower())
      if fConta in string:
         f_value = fMydict[key]['vl_conta']
         break
   return f_value

def transform_Dfp_tri(fDicionario, fConta, fQuarter, fQtdtrimestres, fTrimestres, fTipobalanco, fTipobalanco2):
    f_valor = 0    
    total_tri = fQuarter+3
    f_valor = fDicionario[fTrimestres[fQuarter]][fTipobalanco][fTipobalanco2+' - Demonstração do Resultado'][fConta]['vl_conta']
    dt = datetime.strptime(fTrimestres[fQuarter], '%Y-%m-%d')
    if dt.month == 12:  
        if total_tri > fQtdtrimestres:
            f_valor /= 4
        else:
            f_tri_1 = fDicionario[fTrimestres[fQuarter+1]][fTipobalanco][fTipobalanco2+' - Demonstração do Resultado'][fConta]['vl_conta']
            f_tri_2 = fDicionario[fTrimestres[fQuarter+2]][fTipobalanco][fTipobalanco2+' - Demonstração do Resultado'][fConta]['vl_conta']
            f_tri_3 = fDicionario[fTrimestres[fQuarter+3]][fTipobalanco][fTipobalanco2+' - Demonstração do Resultado'][fConta]['vl_conta']
            f_valor -= (f_tri_1+f_tri_2+f_tri_3) 
    return f_valor

def depreciacao_FDC_tri2(fDicionario, fQuarter, fQtdtrimestres, fTrimestres, fTipobalanco, fTipobalanco2): # deletar após testes
    f_valor = 0    
    f_total_tri = fQuarter+1
    dt = datetime.strptime(fTrimestres[fQuarter], '%Y-%m-%d')
    if f_total_tri+1 >= fQtdtrimestres:
        try:
            f_valor = abs(encontrar_conta(fDicionario[fTrimestres[fQuarter]][fTipobalanco][fTipobalanco2+' - Demonstração do Fluxo de Caixa (Método Indireto)'], 'depreciacoes'))
            if f_valor==0:
                try:
                    f_valor = abs(encontrar_conta(fDicionario[fTrimestres[fQuarter]][fTipobalanco][fTipobalanco2+' - Demonstração do Fluxo de Caixa (Método Indireto)'], 'depreciacao'))
                except KeyError:
                    f_valor = 0   
        except KeyError:
            f_valor = 0
        dt = datetime.strptime(fTrimestres[fQuarter], '%Y-%m-%d')
        if dt.month == 12: f_valor /= 4
        elif dt.month == 9: f_valor /= 3
        elif dt.month == 6: f_valor /= 2
    else:
        if dt.month==3:
            try:
                f_valor = abs(encontrar_conta(fDicionario[fTrimestres[fQuarter]][fTipobalanco][fTipobalanco2+' - Demonstração do Fluxo de Caixa (Método Indireto)'], 'depreciacoes'))
                if f_valor==0:
                    try:
                        f_valor = abs(encontrar_conta(fDicionario[fTrimestres[fQuarter]][fTipobalanco][fTipobalanco2+' - Demonstração do Fluxo de Caixa (Método Indireto)'], 'depreciacao'))
                    except KeyError:
                        f_valor = 0   
            except KeyError:
                f_valor = 0
        else: 
            # trimestre atual
            try:
                f_valor = abs(encontrar_conta(fDicionario[fTrimestres[fQuarter]][fTipobalanco][fTipobalanco2+' - Demonstração do Fluxo de Caixa (Método Indireto)'], 'depreciacoes'))
                if f_valor==0:
                    try:
                        f_valor = abs(encontrar_conta(fDicionario[fTrimestres[fQuarter]][fTipobalanco][fTipobalanco2+' - Demonstração do Fluxo de Caixa (Método Indireto)'], 'depreciacao'))
                    except KeyError:
                        f_valor = 0   
            except KeyError:
                f_valor = 0
            # trimestre anterior
            f_valor_ant = 0
            try:
                f_valor_ant = abs(encontrar_conta(fDicionario[fTrimestres[fQuarter+1]][fTipobalanco][fTipobalanco2+' - Demonstração do Fluxo de Caixa (Método Indireto)'], 'depreciacoes'))
                if f_valor_ant==0:
                    try:
                        f_valor_ant = abs(encontrar_conta(fDicionario[fTrimestres[fQuarter+1]][fTipobalanco][fTipobalanco2+' - Demonstração do Fluxo de Caixa (Método Indireto)'], 'depreciacao'))
                    except KeyError:
                        f_valor_ant = 0   
            except KeyError:
                f_valor_ant = 0
            f_valor = f_valor - f_valor_ant
    return f_valor

def obter_valor_depreciacao(fDicionario, fTrimestre, fTipobalanco, fTipobalanco2):
    chave = fTipobalanco2 + ' - Demonstração do Fluxo de Caixa (Método Indireto)'
    try:
        valor = abs(encontrar_conta(fDicionario[fTrimestre][fTipobalanco][chave], 'depreciacoes'))
        if valor == 0:
            valor = abs(encontrar_conta(fDicionario[fTrimestre][fTipobalanco][chave], 'depreciacao'))
    except KeyError:
        valor = 0
    return valor

def depreciacao_FDC_tri(fDicionario, fQuarter, fQtdtrimestres, fTrimestres, fTipobalanco, fTipobalanco2):
    f_valor = 0    
    f_total_tri = fQuarter + 1
    dt = datetime.strptime(fTrimestres[fQuarter], '%Y-%m-%d')

    if  dt.month == 3:
        f_valor = obter_valor_depreciacao(fDicionario, fTrimestres[fQuarter], fTipobalanco, fTipobalanco2)
    else:
        f_valor_atual = obter_valor_depreciacao(fDicionario, fTrimestres[fQuarter], fTipobalanco, fTipobalanco2)
        f_valor_anterior = obter_valor_depreciacao(fDicionario, fTrimestres[fQuarter + 1], fTipobalanco, fTipobalanco2)
        f_valor = f_valor_atual - f_valor_anterior       

    if f_total_tri + 1 >= fQtdtrimestres:
        divisor = {'12': 4, '9': 3, '6': 2}.get(str(dt.month), 1)
        f_valor /= divisor

    return f_valor

def transform_FDC_tri(fDicionario, fConta, fQuarter, fQtdtrimestres, fTrimestres, fTipobalanco, fTipobalanco2):
    f_total_tri = fQuarter + 1
    dt = datetime.strptime(fTrimestres[fQuarter], '%Y-%m-%d')
    try:
        f_valor = fDicionario[fTrimestres[fQuarter]][fTipobalanco][fTipobalanco2 + ' - Demonstração do Fluxo de Caixa (Método Indireto)'][fConta]['vl_conta']
        if f_total_tri > fQtdtrimestres:
            f_valor /= 4 if dt.month == 12 else 3 if dt.month == 9 else 2 if dt.month == 6 else 1
        elif dt.month != 3:
            f_valor -= fDicionario.get(fTrimestres[f_total_tri], {}).get(fTipobalanco, {}).get(fTipobalanco2 + ' - Demonstração do Fluxo de Caixa (Método Indireto)', {}).get(fConta, {}).get('vl_conta', 0)
    except KeyError:
        try:
            f_valor = fDicionario[fTrimestres[fQuarter]][fTipobalanco][fTipobalanco2 + ' - Demonstração do Fluxo de Caixa (Método Direto)'][fConta]['vl_conta']
        except KeyError:
            f_valor = 0
    return f_valor

# Encontrar empresa balanço por codigo b3
def finddoc_byequity(db, collection_name, b3_code):
    count = 0
    while count != 1:
        collection = db[collection_name]
        documents = collection.find({'equity': b3_code})
        count = collection.count_documents({'equity': b3_code})
        if count > 1:
            print('Lista encontrada: ')
            print(type(documents))
            for doc in documents:
                print(doc['_id'])
            print('CODIGO EXISTENTE EM MAIS DE 01 DOCUMENTO')
            exit()
        if count == 0:
            print("Não foi encontrada a empresa desejada.")
            return None  # Retorna None se não encontrar a empresa
    return documents

# Determinar o tipo de DF avaliar (Consolidado|Individual)
def balance_type(f_empresa, f_trimestre):
    tipo_balanco = 'ERROR'   
    for key in f_empresa[f_trimestre[0]]:
        key2 = re.sub(r'(DF)(Consolidado|Individual)', r'\1 \2', key)       
        try:
             pvalor = f_empresa[f_trimestre[0]][key][key2+' - Demonstração do Resultado']['3.01']['vl_conta']
             if pvalor != 0: 
                 tipo_balanco = key
                 break
        except KeyError:
            print('\nCia nao possui DF CONSOLIDADO')
    novalista_trimestres = []
    if tipo_balanco == 'DFConsolidado':
        for trimestre in f_trimestre:
            if len(f_empresa[trimestre].get(tipo_balanco, {})) != 0:
                    novalista_trimestres.append(trimestre) 
        f_trimestre = novalista_trimestres
    print(f'\nUtilizando balanço: {tipo_balanco}')        
    return tipo_balanco, f_trimestre

def transform_Dfp_tri(f_dicionario, f_conta, f_quarter, f_qtdtrimestres, f_trimestres, f_tipobalanco, f_tipobalanco2):
    f_valor = 0    
    total_tri = f_quarter+3
    f_valor = f_dicionario[f_trimestres[f_quarter]][f_tipobalanco][f_tipobalanco2+' - Demonstração do Resultado'][f_conta]['vl_conta']
    dt = datetime.strptime(f_trimestres[f_quarter], '%Y-%m-%d')
    if dt.month == 12:  
        if total_tri > f_qtdtrimestres:
            f_valor /= 4
        else:
            f_tri_1 = f_dicionario[f_trimestres[f_quarter+1]][f_tipobalanco][f_tipobalanco2+' - Demonstração do Resultado'][f_conta]['vl_conta']
            f_tri_2 = f_dicionario[f_trimestres[f_quarter+2]][f_tipobalanco][f_tipobalanco2+' - Demonstração do Resultado'][f_conta]['vl_conta']
            f_tri_3 = f_dicionario[f_trimestres[f_quarter+3]][f_tipobalanco][f_tipobalanco2+' - Demonstração do Resultado'][f_conta]['vl_conta']
            f_valor -= (f_tri_1+f_tri_2+f_tri_3) 
    return f_valor


# Transformar dados do fluxo de caixa
def transform_FDC_tri(f_dicionario, f_conta, f_quarter, f_qtdtrimestres, f_trimestres, f_tipobalanco, f_tipobalanco2):
    f_valor = 0
    f_total_tri = f_quarter+1
    if f_total_tri > f_qtdtrimestres:
        try:
            f_valor = f_dicionario[f_trimestres[f_quarter]][f_tipobalanco][f_tipobalanco2+' - Demonstração do Fluxo de Caixa (Método Indireto)'][f_conta]['vl_conta']
        except KeyError:
            f_valor = 0
        dt = datetime.strptime(f_trimestres[f_quarter], '%Y-%m-%d')
        if dt.month == 12: f_valor /= 4
        elif dt.month == 9: f_valor /= 3
        elif dt.month == 6: f_valor /= 2
    else:
        # print(f_trimestres[f_quarter], ",", f_conta)
        try:
            f_valor = f_dicionario[f_trimestres[f_quarter]][f_tipobalanco][f_tipobalanco2+' - Demonstração do Fluxo de Caixa (Método Indireto)'][f_conta]['vl_conta']
        except KeyError:
            try:
                f_valor = f_dicionario[f_trimestres[f_quarter]][f_tipobalanco][f_tipobalanco2+' - Demonstração do Fluxo de Caixa (Método Direto)'][f_conta]['vl_conta']
            except KeyError:
                f_valor = 0
        dt = datetime.strptime(f_trimestres[f_quarter], '%Y-%m-%d')
        if dt.month != 3:
            try:
                f_valor -= f_dicionario[f_trimestres[f_total_tri]][f_tipobalanco][f_tipobalanco2+' - Demonstração do Fluxo de Caixa (Método Indireto)'][f_conta]['vl_conta']
            except KeyError:
                f_valor = 0
    return f_valor


# Transformar dados do fdc / depreciacao
def depreciacao_FDC_tri(f_dicionario, f_quarter, f_qtdtrimestres, f_trimestres, f_tipobalanco, f_tipobalanco2):
    f_valor = 0    
    f_total_tri = f_quarter+1
    dt = datetime.strptime(f_trimestres[f_quarter], '%Y-%m-%d')
    if f_total_tri+1 >= f_qtdtrimestres:
        try:
            f_valor = abs(encontrar_conta(f_dicionario[f_trimestres[f_quarter]][f_tipobalanco][f_tipobalanco2+' - Demonstração do Fluxo de Caixa (Método Indireto)'], 'depreciacoes'))
            if f_valor==0:
                try:
                    f_valor = abs(encontrar_conta(f_dicionario[f_trimestres[f_quarter]][f_tipobalanco][f_tipobalanco2+' - Demonstração do Fluxo de Caixa (Método Indireto)'], 'depreciacao'))
                except KeyError:
                    f_valor = 0   
        except KeyError:
            f_valor = 0
        dt = datetime.strptime(f_trimestres[f_quarter], '%Y-%m-%d')
        if dt.month == 12: f_valor /= 4
        elif dt.month == 9: f_valor /= 3
        elif dt.month == 6: f_valor /= 2
    else:
        if dt.month==3:
            try:
                f_valor = abs(encontrar_conta(f_dicionario[f_trimestres[f_quarter]][f_tipobalanco][f_tipobalanco2+' - Demonstração do Fluxo de Caixa (Método Indireto)'], 'depreciacoes'))
                if f_valor==0:
                    try:
                        f_valor = abs(encontrar_conta(f_dicionario[f_trimestres[f_quarter]][f_tipobalanco][f_tipobalanco2+' - Demonstração do Fluxo de Caixa (Método Indireto)'], 'depreciacao'))
                    except KeyError:
                        f_valor = 0   
            except KeyError:
                f_valor = 0
        else: 
            # trimestre atual
            try:
                f_valor = abs(encontrar_conta(f_dicionario[f_trimestres[f_quarter]][f_tipobalanco][f_tipobalanco2+' - Demonstração do Fluxo de Caixa (Método Indireto)'], 'depreciacoes'))
                if f_valor==0:
                    try:
                        f_valor = abs(encontrar_conta(f_dicionario[f_trimestres[f_quarter]][f_tipobalanco][f_tipobalanco2+' - Demonstração do Fluxo de Caixa (Método Indireto)'], 'depreciacao'))
                    except KeyError:
                        f_valor = 0   
            except KeyError:
                f_valor = 0
            # trimestre anterior
            f_valor_ant = 0
            try:
                f_valor_ant = abs(encontrar_conta(f_dicionario[f_trimestres[f_quarter+1]][f_tipobalanco][f_tipobalanco2+' - Demonstração do Fluxo de Caixa (Método Indireto)'], 'depreciacoes'))
                if f_valor_ant==0:
                    try:
                        f_valor_ant = abs(encontrar_conta(f_dicionario[f_trimestres[f_quarter+1]][f_tipobalanco][f_tipobalanco2+' - Demonstração do Fluxo de Caixa (Método Indireto)'], 'depreciacao'))
                    except KeyError:
                        f_valor_ant = 0   
            except KeyError:
                f_valor_ant = 0
            f_valor = f_valor - f_valor_ant
    return f_valor

# Encontrar contas depois de codigos
def encontrar_conta(f_mydict, conta):
   f_value = 0 
   for key in f_mydict:
      string = unidecode(f_mydict[key]['ds_conta'].lower())
      if conta in string:
         f_value = f_mydict[key]['vl_conta']
         break
   return f_value

# Definição de nome do arquivo da planilha
def nome_arquivo(f_b3issuercode, f_trimestre):
    f_data = datetime.strptime(f_trimestre, "%Y-%m-%d")
    f_mes = f_data.month    
    if f_mes == 3:
        f_name = '1Q'+f_trimestre[2]+f_trimestre[3]
    elif f_mes == 6:
        f_name = '2Q'+f_trimestre[2]+f_trimestre[3]
    elif f_mes == 9:
        f_name = '3Q'+f_trimestre[2]+f_trimestre[3]
    elif f_mes == 12:
        f_name = '4Q'+f_trimestre[2]+f_trimestre[3]
    else:
        f_name = 'ERROR'
    return (f_b3issuercode+' '+f_name)


# stock = ['RADL3', 'SBSP3', 'PETR4', 'VALE3']

stock = [
     'VALE3', 'PETR4',  'ELET3',  'ABEV3', 'RENT3', 'ITSA4',  # 'B3SA3', 'ITUB4','BBDC4', 'BBAS3',
     'WEGE3', 'BPAC11','EQTL3', 'PRIO3', 'SUZB3', 'RADL3', 'RAIL3', 'GGBR4', 'UGPA3', #'RDOR3',
     'JBSS3', 'VBBR3', 'BRFS3',   'VIVT3', 'CSAN3', 'ENEV3', 'HAPV3', 'CMIG4', 'SBSP3', #'BBSE3',
     'TOTS3', 'KLBN11','CPLE6', 'ENGI11','EMBR3', 'HYPE3', 'LREN3', #' TIMS3', #'ALOS3', 'ASAI3', 
     'CCRO3'] 

stock = [ 'EGIE3', 'CSNA3', #'CMIN3', #'GOAU4', # 'SANB11','TAEE11','RRRP3', 
     'MULT3', 'CPFE3', 'MGLU3', 'FLRY3', 'YDUQ3', 'AZUL4', 'CRFB3', 'CYRE3', 'COGN3', # 'BRAP4', 
     'BRKM5',  'IGTI11', 'CIEL3',  'MRVE3', 'SLCE3', 'USIM5', # 'RECV3','SMTO3','RAIZ4','VAMO3',
     'ARZZ3', 'MRFG3', 'DXCO3', 'ALPA4', 'BEEF3', 'GOLL4', 'EZTC3', # 'SOMA3','IRBR3',  'LWSA3', 
     'CVCB3', 'PCAR3'] # 'PETZ3','BHIA3',

print(len(stock))

for y in range(len(stock)):
    print(y)
    
    info_input = [] 
    info_input.append(stock[y])
    document = finddoc_byequity(db, collection, stock[y])
    doc_stock = document.next() 
    id_empresa = doc_stock['_id']; print(id_empresa)
    tipo_balanco1, tipo_balanco2, trimestres_lista = upload_infofin(doc_stock, doc_stock) 

    # ebit - calculo 
    ebit = []
    for i in range(len(trimestres_lista)):
        ebit.append(transform_Dfp_tri(doc_stock, '3.05', i, len(trimestres_lista), trimestres_lista, tipo_balanco1, tipo_balanco2))
    ebit = np.array(ebit)
    info_input.append(round(ebit[0]/ebit[0+4], 2))
    info_input.append(round(ebit[1]/ebit[1+4], 2))
    info_input.append(round(ebit[2]/ebit[2+4], 2)) 
    info_input.append(round(info_input[2] - info_input[1], 2))
    info_input.append(round(info_input[3] - info_input[2], 2))
    info_input.append(round(np.sum(ebit[0:4])/np.sum(ebit[4:4+4]), 2)) # 6
    info_input.append(round(np.sum(ebit[1:1+4])/np.sum(ebit[5:5+4]), 2)) # 7
    info_input.append(round(np.sum(ebit[2:2+4])/np.sum(ebit[6:6+4]), 2)) # 8
    info_input.append(round(info_input[7] - info_input[8], 2)) # 9 
    info_input.append(round(info_input[6] - info_input[7], 2)) # 10

    # ebitda - calculo
    ebitda = []
    for i in range(len(trimestres_lista)):
        ebit_value = transform_Dfp_tri(doc_stock, '3.05', i, len(trimestres_lista), trimestres_lista, tipo_balanco1, tipo_balanco2) 
        depreciation_value = depreciacao_FDC_tri(doc_stock, i, len(trimestres_lista), trimestres_lista, tipo_balanco1, tipo_balanco2) 
        ebitda.append(ebit_value+depreciation_value)
    ebitda = np.array(ebitda)
    info_input.append(round(ebitda[0]/ebitda[0+4], 2)) # 11
    info_input.append(round(ebitda[1]/ebitda[1+4], 2)) # 12 
    info_input.append(round(ebitda[2]/ebitda[2+4], 2)) # 13
    info_input.append(round(info_input[12] - info_input[13], 2)) # 14
    info_input.append(round(info_input[11] - info_input[12], 2)) # 15
    info_input.append(round(np.sum(ebitda[0:4])/np.sum(ebitda[4:4+4]), 2)) # 16
    info_input.append(round(np.sum(ebitda[1:1+4])/np.sum(ebitda[5:5+4]), 2)) # 17
    info_input.append(round(np.sum(ebitda[2:2+4])/np.sum(ebitda[6:6+4]), 2)) # 18
    info_input.append(round(info_input[17] - info_input[18], 2)) # 19
    info_input.append(round(info_input[16] - info_input[17], 2)) # 20

    # fluxo de caixa operacional - calculo
    fluxocaixa_operacional = []
    for i in range(len(trimestres_lista)):
        fluxocaixa_value = transform_FDC_tri(doc_stock, '6.01.01', i, len(trimestres_lista), trimestres_lista, tipo_balanco1, tipo_balanco2)
        fluxocaixa_operacional.append(fluxocaixa_value)
    fluxocaixa_operacional = np.array(fluxocaixa_operacional)
    info_input.append(round(fluxocaixa_operacional[0]/fluxocaixa_operacional[0+4], 2)) # 21
    info_input.append(round(fluxocaixa_operacional[1]/fluxocaixa_operacional[1+4], 2)) # 22
    info_input.append(round(fluxocaixa_operacional[2]/fluxocaixa_operacional[2+4], 2)) # 23
    info_input.append(round(info_input[22] - info_input[23], 2)) # 24
    info_input.append(round(info_input[21] - info_input[22], 2)) # 25
    info_input.append(round(np.sum(fluxocaixa_operacional[0:4])/np.sum(fluxocaixa_operacional[4:4+4]), 2)) # 26
    info_input.append(round(np.sum(fluxocaixa_operacional[1:1+4])/np.sum(fluxocaixa_operacional[5:5+4]), 2)) # 27
    info_input.append(round(np.sum(fluxocaixa_operacional[2:2+4])/np.sum(fluxocaixa_operacional[6:6+4]), 2)) # 28
    info_input.append(round(info_input[27] - info_input[28], 2)) # 29
    info_input.append(round(info_input[26] - info_input[27], 2)) # 30

    # datas 
    document_inforesultados = finddoc_byname(db, 'divulgacaoinfo_resultados', id_empresa, '_30/06/2023').next()
    nome_cia = document_inforesultados['nome_empresa']
    hora_envioresultado = document_inforesultados['hora_envio']

    info_input.append(fConverterstr_data(hora_envioresultado) + timedelta(fExtrair_hora(hora_envioresultado))) # '2023-08-10' # 31
    info_input.append('2023-10-10') # 32
    info_input.append(fGet_adjclosing_price(df_stocks, str(info_input[31]), stock[y])) # 33
    info_input.append(fGet_adjclosing_price(df_stocks, info_input[32], stock[y])) # 34
    info_input.append(round(info_input[34]/ info_input[33]- 1, 4)) # 35

    preco_inicial = fGet_adjclosing_price(df_stocks, str(info_input[31]), 'BOVA11')
    preco_final = fGet_adjclosing_price(df_stocks, info_input[32], 'BOVA11')
    info_input.append(round(preco_final/ preco_inicial- 1, 4)) # 36

    preco_inicial = fGet_adjclosing_price(df_stocks, str(info_input[31]), 'SMAL11')
    preco_final = fGet_adjclosing_price(df_stocks, info_input[32], 'SMAL11')
    info_input.append(round(preco_final/ preco_inicial- 1, 4)) # 37


    # manipulação arquivo excel
    for linha in range(1, planilha.max_row + 1):
        if not planilha[f'A{linha}'].value:
            linha  # Retorna o número da linha da primeira célula em branco

    # print(f'Linha: {linha+1} e ativo {stock[y]}')
    for w in range(1, len(info_input)+1):
        planilha.cell(linha+1, w).value = info_input[w-1]

arquivo.save('estquant_fundamentus_2.xlsx')
arquivo.close()
