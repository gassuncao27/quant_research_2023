from pymongo import MongoClient
from pymongo.cursor import Cursor
from datetime import datetime, timedelta
from unidecode import unidecode
from dotenv import load_dotenv

import openpyxl

# from openpyxl.styles import Font
# from openpyxl.styles import PatternFill
# from openpyxl import load_workbook

import pandas as pd

import numpy as np
import os
import re

# Acessando o banco MongoDB Compass
cliente = MongoClient(
    "ddns.cartor.com.br",
    port=57017,
    username='cartor',
    password='Cartor1212',
    authSource='admin',
    authMechanism='DEFAULT'
)

db = cliente.CVM
collection = 'Empresas_fin6'

load_dotenv()   

df_stocks = pd.read_csv('stockprices_dashboard.txt', header=None)
df_stocks.columns = ['Stock', 'Date', 'Price']
df_stocks['Date'] = pd.to_datetime(df_stocks['Date'])
df_stocks.set_index('Date', inplace=True)

# stocks_path = os.getenv('AMBIENTE_MAC')+'stocks_data.parquet'
# df_stocks = pd.read_parquet(stocks_path); df_stocks.fillna(0, inplace=True)

# abertura template excel 
arquivo = openpyxl.load_workbook('estquant_fundamentus.xlsx')
planilha = arquivo['Planilha1']


def get_adjclosing_price(df, date, ticker):
    if isinstance(date, str):
        date = pd.to_datetime(date)
    filtered_df = df[(df['data'] == date) & (df['ticker'] == ticker)]
    if not filtered_df.empty:
        return filtered_df['preco_fechamento_ajustado'].iloc[0]
    else:
        return None
    
def get_closingprice(df, date, stock):
    try:
        data_formatada = pd.to_datetime(date)
        resultado = df[(df.index == data_formatada) & (df['Stock'] == stock)]
        if not resultado.empty:
            return resultado['Price'].iloc[0]
        else:
            return "Preço não encontrado para a data e stock fornecidos."
    except Exception as e:
        return f"Ocorreu um erro: {e}"

    
def encontrar_close(df, stock_name, fdata_divulgacao):
    try:
        filtro = (df['Stock'] == stock_name) & (df['Date'] == fdata_divulgacao)
        close_value = df.loc[filtro, 'Close'].values[0]
        return close_value
    except IndexError:
        try:
            filtro = (df['Stock'] == stock_name) & (df['Date'] == (fdata_divulgacao+timedelta(days=1)))
            close_value = df.loc[filtro, 'Close'].values[0]
            return close_value            
        except IndexError:        
            try:
                filtro = (df['Stock'] == stock_name) & (df['Date'] == (fdata_divulgacao+timedelta(days=2)))
                close_value = df.loc[filtro, 'Close'].values[0]
                return close_value
            except IndexError:
                try:
                    filtro = (df['Stock'] == stock_name) & (df['Date'] == (fdata_divulgacao+timedelta(days=2)))
                    close_value = df.loc[filtro, 'Close'].values[0]
                    return close_value
                except IndexError:                                                                
                    return None 
    
def converterstr_data(fData_hora_string):
    data_hora = datetime.strptime(fData_hora_string, "%d/%m/%Y %H:%M:%S")
    data = data_hora.date()
    return data    

def extrair_hora(fData_hora_str):
    data_hora = datetime.strptime(fData_hora_str, "%d/%m/%Y %H:%M:%S")
    horas = data_hora.hour
    if horas > 17:
        aumentar_dias = 1
    else:
        aumentar_dias = 0
    return aumentar_dias

def finddoc_byequity(fDb, fCollection_name, fEquity): # acha o documento da empresa pelo ticker da bolsa 
    count = 0
    while count != 1:
        collection = fDb[fCollection_name]
        documents = collection.find({'equity': fEquity})
        count = collection.count_documents({'equity': fEquity})
        if count > 1:
            print('Lista encontrada: ')
            print(type(documents))
            for doc in documents:
                print(doc['_id'])
            print('CODIGO EXISTENTE EM MAIS DE 01 DOCUMENTO')
            exit()
        if count == 0:
            print("Não foi encontrada a empresa desejada.")
            return None  # Retorna None se não encontrar a empresa
    return documents

def finddoc_byname(fDb, fCollection_name, fContaining_word, fTrimestre=False):
    count = 0
    while count != 1:
        if fTrimestre:
            collection = fDb[fCollection_name]
            regex = re.compile(fContaining_word+fTrimestre, re.IGNORECASE)
            documents = collection.find({'_id': regex})
            count = collection.count_documents({'_id': regex})
        else:
            collection = fDb[fCollection_name]
            regex = re.compile(fContaining_word, re.IGNORECASE)
            documents = collection.find({'_id': regex})
            count = collection.count_documents({'_id': regex})
        if count > 1:
            print(type(documents))
            print('erro aconteceu - checar base')
            for doc in documents:
                print(doc['_id'])
            break
        if count == 0:
            print("Não foi encontrada a empresa desejada.")
            return None
    return documents

def finddoc_by_codcvm(fDb, fCollection_name, fCodcvm, fContaining_word, fTrimestre): 
    count = 0
    while count != 1:
        collection = fDb[fCollection_name]
        documents = collection.find({'cod_cvm': fCodcvm})
        count = collection.count_documents({'cod_cvm': fCodcvm})
        if count > 1:
            for doc in documents:
                if fTrimestre in doc['_id']:
                    documents = doc
                    count = 1
                    break
        if count == 0:
            print("Não foi encontrada a empresa desejada.")
            return None      
    if isinstance(documents, Cursor):
        return documents.next()
    else: 
        return documents

def upload_infofin(fData_fin): # retorna os trimestres da empresa 
    trimestres_upload = []
    for chave, valor in fData_fin.items():
        if isinstance(chave, str) and chave.startswith('2'):
            trimestres_upload.append(chave)
    trimestres_upload = sorted(trimestres_upload, reverse=True)     
    tipo_balanco, trimestres_upload = get_balance_type(fData_fin, trimestres_upload)
    if tipo_balanco == 'ERROR': 
        print('Erro na captura do balanço'); sys.exit()
    tipo_balanco2 = re.sub(r'(DF)(Consolidado|Individual)', r'\1 \2', tipo_balanco)
    return tipo_balanco, tipo_balanco2, trimestres_upload

def get_balance_type(fEmpresa, fTrimestre): # seleciona o tipo de balanco existente
    tipo_balanco = 'ERROR'
    for key in fEmpresa[fTrimestre[0]]:
        key2 = re.sub(r'(DF)(Consolidado|Individual)', r'\1 \2', key)       
        try:
            nr_trimestre = 0
            while nr_trimestre < 8:
                pvalor = fEmpresa[fTrimestre[nr_trimestre]][key][key2+' - Demonstração do Resultado']['3.05']['vl_conta']
                if pvalor !=0: 
                    nr_trimestre+=1
                else:
                    break
            if nr_trimestre == 8:
                tipo_balanco = key
        except KeyError:
            print(f'Empresa não possui {key} suficiente')
        if tipo_balanco != 'ERROR': break
    novalista_trimestres = []
    for i, trimestre in enumerate(fTrimestre):        
        if len(fEmpresa[trimestre].get(tipo_balanco, {})) != 0:
                novalista_trimestres.append(trimestre) 
    fTrimestre = novalista_trimestres
    return tipo_balanco, fTrimestre

def encontrar_conta(fMydict, fConta):
   f_value = 0 
   for key in fMydict:
      string = unidecode(fMydict[key]['ds_conta'].lower())
      if fConta in string:
         f_value = fMydict[key]['vl_conta']
         break
   return f_value

def transform_Dfp_tri(fDicionario, fConta, fQuarter, fQtdtrimestres, fTrimestres, fTipobalanco, fTipobalanco2):
    f_valor = 0    
    total_tri = fQuarter+3
    f_valor = fDicionario[fTrimestres[fQuarter]][fTipobalanco][fTipobalanco2+' - Demonstração do Resultado'][fConta]['vl_conta']
    dt = datetime.strptime(fTrimestres[fQuarter], '%Y-%m-%d')
    if dt.month == 12:  
        if total_tri > fQtdtrimestres:
            f_valor /= 4
        else:
            f_tri_1 = fDicionario[fTrimestres[fQuarter+1]][fTipobalanco][fTipobalanco2+' - Demonstração do Resultado'][fConta]['vl_conta']
            f_tri_2 = fDicionario[fTrimestres[fQuarter+2]][fTipobalanco][fTipobalanco2+' - Demonstração do Resultado'][fConta]['vl_conta']
            f_tri_3 = fDicionario[fTrimestres[fQuarter+3]][fTipobalanco][fTipobalanco2+' - Demonstração do Resultado'][fConta]['vl_conta']
            f_valor -= (f_tri_1+f_tri_2+f_tri_3) 
    return f_valor

def obter_valor_depreciacao(fDicionario, fTrimestre, fTipobalanco, fTipobalanco2):
    chave = fTipobalanco2 + ' - Demonstração do Fluxo de Caixa (Método Indireto)'
    try:
        valor = abs(encontrar_conta(fDicionario[fTrimestre][fTipobalanco][chave], 'depreciacoes'))
        if valor == 0:
            valor = abs(encontrar_conta(fDicionario[fTrimestre][fTipobalanco][chave], 'depreciacao'))
    except KeyError:
        valor = 0
    return valor

def depreciacao_FDC_tri(fDicionario, fQuarter, fQtdtrimestres, fTrimestres, fTipobalanco, fTipobalanco2):
    f_valor = 0    
    f_total_tri = fQuarter + 1
    dt = datetime.strptime(fTrimestres[fQuarter], '%Y-%m-%d')

    if  dt.month == 3:
        f_valor = obter_valor_depreciacao(fDicionario, fTrimestres[fQuarter], fTipobalanco, fTipobalanco2)
    elif f_total_tri + 1 >= fQtdtrimestres:
        divisor = {'12': 4, '9': 3, '6': 2}.get(str(dt.month), 1)
        f_valor /= divisor
    else:
        f_valor_atual = obter_valor_depreciacao(fDicionario, fTrimestres[fQuarter], fTipobalanco, fTipobalanco2)
        f_valor_anterior = obter_valor_depreciacao(fDicionario, fTrimestres[fQuarter + 1], fTipobalanco, fTipobalanco2) # tratar
        f_valor = f_valor_atual - f_valor_anterior       
    return f_valor

def transform_FDC_tri(fDicionario, fConta, fQuarter, fQtdtrimestres, fTrimestres, fTipobalanco, fTipobalanco2):
    f_total_tri = fQuarter + 1
    dt = datetime.strptime(fTrimestres[fQuarter], '%Y-%m-%d')
    f_valor = fDicionario[fTrimestres[fQuarter]][fTipobalanco][fTipobalanco2 + ' - Demonstração do Fluxo de Caixa (Método Indireto)'][fConta]['vl_conta']
    if dt.month == 3:
        return f_valor
    elif f_total_tri < len(fTrimestres):
        f_valor -= fDicionario.get(fTrimestres[f_total_tri], {}).get(fTipobalanco, {}).get(fTipobalanco2 + ' - Demonstração do Fluxo de Caixa (Método Indireto)', {}).get(fConta, {}).get('vl_conta', 0)
    else:
        if dt.month == 12:
            f_valor /= 4 
        elif dt.month == 9:
            f_valor /= 3 
        else:
            f_valor /= 2
    if f_valor == 0: f_valor = 1
    return f_valor


# stock = ['AESB3']

# stock = [
#     'VALE3', 'PETR4', 'ELET3', 'ABEV3', 'RENT3', 'ITSA4', 'B3SA3', 'WEGE3', 'BPAC11','EQTL3',
#     'PRIO3', 'SUZB3', 'RADL3', 'RAIL3', 'GGBR4', 'UGPA3', 'RDOR3', 'JBSS3', 'VBBR3', 'BRFS3',
#     'VIVT3', 'CSAN3', 'ENEV3', 'HAPV3', 'CMIG4', 'SBSP3', 'TOTS3', 'KLBN11','CPLE6','ENGI11',
#     'EMBR3', 'HYPE3', 'LREN3', 'TIMS3', 'ASAI3', 'CCRO3', 'EGIE3', 'CSNA3', 'CMIN3', 'GOAU4', 
#     'TAEE11','RRRP3', 'MULT3', 'CPFE3', 'MGLU3', 'FLRY3', 'YDUQ3', 'AZUL4', 'CRFB3', 'CYRE3', 
#     'COGN3', 'BRAP4', 'BRKM5','IGTI11', 'CIEL3', "TTEN3", "ABCB4", "AERI3", "AALR3", "ALPA4",
#     "ALUP11","AMBP3", "ANIM3", "ARZZ3", "ARML3", "BMOB3", "CEAB3", "CLSA3", "CSMG3", "CURY3", 
#     "CVCB3", "DASA3", "DXCO3", "PNVL3", "DIRR3", "ECOR3", "ENAT3", "ESPA3", "EVEN3", "EZTC3", 
#     "FESA4", "FRAS3", "GFSA3", "GOLL4", "GGPS3", "GRND3", "SBFG3", "GUAR3", "HBSA3", "INTB3",
#     "MYPK3", "RANI3", "JHSF3", "KEPL3", "LAVV3", "LWSA3", "LOGG3", "LOGN3", "POMO4", "MRFG3", 
#     "MATD3", "LEVE3", "BEEF3", "MOVI3", "MRVE3", "MLAS3", "ODPV3", "ONCO3", "ORVR3", "PCAR3", 
#     "PGMN3", "RECV3", "PETZ3", "PLPL3", "PTBL3", "POSI3", "QUAL3", "LJQQ3", "RAPT4", "ROMI3",
#     "SAPR11","STBP3", "SEQL3", "SEER3", "SIMH3", "SLCE3", "SMFT3", "TASA4", "TGMA3", "TEND3", 
#     "TRIS3", "TUPY3", "UNIP6", "USIM5", "VLID3", "VULC3", "WIZC3", "ZAMP3" ]

stock = ['VALE3', 'PETR4']

for y in range(len(stock)):
    print(stock[y])
    
    info_input = [] 
    info_input.append(stock[y])
    document = finddoc_byequity(db, collection, stock[y])
    doc_stock = document.next() 
    id_empresa = doc_stock['_id']; 
    tipo_balanco1, tipo_balanco2, trimestres_lista = upload_infofin(doc_stock) 

    # ebit - calculo 
    lucroliquido = []
    for i in range(len(trimestres_lista)):
        lucroliquido.append(transform_Dfp_tri(doc_stock, '3.09', i, len(trimestres_lista), trimestres_lista, tipo_balanco1, tipo_balanco2) )
    lucroliquido = np.array(lucroliquido)
    info_input.append(round(lucroliquido[0]/lucroliquido[0+4], 2))
    info_input.append(round(lucroliquido[1]/lucroliquido[1+4], 2))
    info_input.append(round(lucroliquido[2]/lucroliquido[2+4], 2)) 
    info_input.append(round(info_input[2] - info_input[1], 2))
    info_input.append(round(info_input[3] - info_input[2], 2))
    info_input.append(round(np.sum(lucroliquido[0:4])/np.sum(lucroliquido[4:4+4]), 2)) # 6
    info_input.append(round(np.sum(lucroliquido[1:1+4])/np.sum(lucroliquido[5:5+4]), 2)) # 7
    info_input.append(round(np.sum(lucroliquido[2:2+4])/np.sum(lucroliquido[6:6+4]), 2)) # 8
    info_input.append(round(info_input[7] - info_input[8], 2)) # 9 
    info_input.append(round(info_input[6] - info_input[7], 2)) # 10

    # ebitda - calculo
    ebitda = []
    for i in range(len(trimestres_lista)):
        ebit_value = transform_Dfp_tri(doc_stock, '3.05', i, len(trimestres_lista), trimestres_lista, tipo_balanco1, tipo_balanco2) 
        depreciation_value = depreciacao_FDC_tri(doc_stock, i, len(trimestres_lista), trimestres_lista, tipo_balanco1, tipo_balanco2) 
        ebitda.append(ebit_value+depreciation_value)
    ebitda = np.array(ebitda)
    info_input.append(round(ebitda[0]/ebitda[0+4], 2)) # 11
    info_input.append(round(ebitda[1]/ebitda[1+4], 2)) # 12 
    info_input.append(round(ebitda[2]/ebitda[2+4], 2)) # 13
    info_input.append(round(info_input[12] - info_input[13], 2)) # 14
    info_input.append(round(info_input[11] - info_input[12], 2)) # 15
    info_input.append(round(np.sum(ebitda[0:4])/np.sum(ebitda[4:4+4]), 2)) # 16
    info_input.append(round(np.sum(ebitda[1:1+4])/np.sum(ebitda[5:5+4]), 2)) # 17
    info_input.append(round(np.sum(ebitda[2:2+4])/np.sum(ebitda[6:6+4]), 2)) # 18
    info_input.append(round(info_input[17] - info_input[18], 2)) # 19
    info_input.append(round(info_input[16] - info_input[17], 2)) # 20

    # fluxo de caixa operacional - calculo
    fluxocaixa_operacional = []
    for i in range(len(trimestres_lista)):
        fluxocaixa_value = transform_FDC_tri(doc_stock, '6.01.01', i, len(trimestres_lista), trimestres_lista, tipo_balanco1, tipo_balanco2)
        fluxocaixa_operacional.append(fluxocaixa_value)
    fluxocaixa_operacional = np.array(fluxocaixa_operacional)
    info_input.append(round(fluxocaixa_operacional[0]/fluxocaixa_operacional[0+4], 2)) # 21
    info_input.append(round(fluxocaixa_operacional[1]/fluxocaixa_operacional[1+4], 2)) # 22
    info_input.append(round(fluxocaixa_operacional[2]/fluxocaixa_operacional[2+4], 2)) # 23
    info_input.append(round(info_input[22] - info_input[23], 2)) # 24
    info_input.append(round(info_input[21] - info_input[22], 2)) # 25
    info_input.append(round(np.sum(fluxocaixa_operacional[0:4])/np.sum(fluxocaixa_operacional[4:4+4]), 2)) # 26
    info_input.append(round(np.sum(fluxocaixa_operacional[1:1+4])/np.sum(fluxocaixa_operacional[5:5+4]), 2)) # 27
    info_input.append(round(np.sum(fluxocaixa_operacional[2:2+4])/np.sum(fluxocaixa_operacional[6:6+4]), 2)) # 28
    info_input.append(round(info_input[27] - info_input[28], 2)) # 29
    info_input.append(round(info_input[26] - info_input[27], 2)) # 30

    # datas 
    document_inforesultados = finddoc_by_codcvm(db, 'divulgacaoinfo_resultados', doc_stock['cod_cvm'], id_empresa, '_30/06/2023')
    nome_cia = document_inforesultados['nome_empresa']
    hora_envioresultado = document_inforesultados['hora_envio']

    # precos segundo trimestre
    info_input.append('2023-08-10') # 31    
    info_input.append('2023-10-10') # 32
    info_input.append(get_closingprice(df_stocks, info_input[31], stock[y])) # 33
    info_input.append(get_closingprice(df_stocks, info_input[32], stock[y])) # 34
    info_input.append(round(info_input[34]/ info_input[33]- 1, 4)) # 35

    preco_inicial = get_closingprice(df_stocks, info_input[31], 'BOVA11') 
    preco_final = get_closingprice(df_stocks, info_input[32], 'BOVA11')
    info_input.append(round(preco_final/ preco_inicial- 1, 4)) # 36

    preco_inicial = get_closingprice(df_stocks, info_input[31], 'SMAL11') 
    preco_final = get_closingprice(df_stocks, info_input[32], 'SMAL11')    
    info_input.append(round(preco_final/ preco_inicial- 1, 4)) # 37

    # adicionando preços do 3trimestre
    info_input.append('2023-10-11') # 38    
    info_input.append('2023-12-01') # 39
    info_input.append(get_closingprice(df_stocks, info_input[38], stock[y])) # 40
    info_input.append(get_closingprice(df_stocks, info_input[39], stock[y])) # 41
    info_input.append(round(info_input[41]/ info_input[40]- 1, 4)) # 42  

    preco_inicial = get_closingprice(df_stocks, info_input[38], 'BOVA11') 
    preco_final = get_closingprice(df_stocks, info_input[39], 'BOVA11')
    info_input.append(round(preco_final/ preco_inicial- 1, 4)) # 43

    preco_inicial = get_closingprice(df_stocks, info_input[38], 'SMAL11') 
    preco_final = get_closingprice(df_stocks, info_input[39], 'SMAL11')    
    info_input.append(round(preco_final/ preco_inicial- 1, 4)) # 44  

    # manipulação arquivo excel
    for linha in range(1, planilha.max_row + 1):
        if not planilha[f'A{linha}'].value:
            linha  # Retorna o número da linha da primeira célula em branco

    # print(f'Linha: {linha+1} e ativo {stock[y]}')
    for w in range(1, len(info_input)+1):
        planilha.cell(linha+1, w).value = info_input[w-1]

arquivo.save('estquant_fundamentus_ll.xlsx')
arquivo.close()
